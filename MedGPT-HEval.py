
###############################################################################################################################################

#---------------------------------------------------------------- MedGPT-HEval ---------------------------------------------------------------#

###############################################################################################################################################


import pyperclip
import json
import time
import re
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from seleniumbase import Driver
import numpy as np
from transformers import BartTokenizer, BartForConditionalGeneration
import torch
from sentence_transformers import SentenceTransformer, util

json_input_path = r"C:\Users\XXXXXXXXXX\Healthcare GPT Assessment\Open-source_responses_hallu.json"
excel_path = r"C:\Users\XXXXXXXXX\Healthcare GPT Assessment\gpt_metadata_output_MedLLM_processed_0.xlsx"
sheet_name = "Open Source LLM"
chatgpt_url = "https://chatgpt.com/"
output_start_col = 8
#start_row = 2
#max_row = 20
response_keys = ["response_1", "response_2", "response_3", "response_4", "response_5"]
WAIT_BETWEEN_RESPONSES = 20

PROMPT_LIMIT = 160
TIME_WINDOW = 3600  

context_text = (
    "Birth trauma is a known complication in neonates with macrosomia, especially following prolonged or difficult labor. "
    "Clavicle fractures are one of the most common skeletal injuries in this setting and are often recognized by localized "
    "tenderness, crepitus, and asymmetry in neonatal reflexes such as the Moro reflex."
)

scoring_rubric_prompt = """You are a medical evaluation assistant. Please evaluate the following model response according to the five hallucination criteria. 
Assign a score between 0 and 1 for each criterion. Only return the scores, nothing else.

Criteria:
1. Consistency: Is the answer aligned with the question and context? 
2. Factual Accuracy: Are all clinical claims accurate and medically sound?
3. Completeness: Does the answer cover all relevant points?
4. Citation Reliability: Are cited sources real and verifiable? (or absent)
5. Inference Justification: Are conclusions logical and evidence-based?

Evaluation Steps:
1. Check that the context is complete and coherent
2. Ensure that the model output logically addresses the question using relevant elements from the context
3. Fact-check all clinical claims or statistics using reputable sources (e.g., PubMed, WHO, CDC, clinical guidelines)
4. Verify that all cited sources (if any) are real and not hallucinated
5. Determine if the conclusion logically follows from the information provided

Please return only your scores in the following format:
Consistency: X.XX  
Factual Accuracy: X.XX  
Completeness: X.XX  
Citation Reliability: X.XX  
Inference Justification: X.XX  
Final Score (average): X.XX"""

tokenizer = BartTokenizer.from_pretrained("facebook/bart-large-cnn")
model = BartForConditionalGeneration.from_pretrained("facebook/bart-large-cnn")
model.eval()
embedder = SentenceTransformer('pritamdeka/BioBERT-mnli-snli-scinli-scitail-mednli-stsb')

def compute_bartscore(reference: str, hypothesis: str) -> float:
    inputs = tokenizer(hypothesis, return_tensors="pt", max_length=1024, truncation=True)
    with torch.no_grad():
        outputs = model(**inputs, labels=tokenizer(reference, return_tensors="pt").input_ids)
        loss = outputs.loss.item()
    return -loss

def compute_semantic_entropy(text: str) -> float:
    inputs = tokenizer(text, return_tensors="pt", max_length=1024, truncation=True)
    with torch.no_grad():
        outputs = model(**inputs)
        logits = outputs.logits
    probs = torch.nn.functional.softmax(logits, dim=-1)
    log_probs = torch.nn.functional.log_softmax(logits, dim=-1)
    entropy = -torch.sum(probs * log_probs, dim=-1)
    avg_entropy = torch.mean(entropy).item()
    return avg_entropy

def compute_cosine_similarity(reference: str, hypothesis: str) -> float:
    embeddings = embedder.encode([reference, hypothesis], convert_to_tensor=True)
    similarity = util.pytorch_cos_sim(embeddings[0], embeddings[1])
    return similarity.item()

def format_prompt(question, response):
    return (
        f"{scoring_rubric_prompt.strip()}\n\n"
        f"=== CONTEXT ===\n{context_text.strip()}\n\n"
        f"=== QUESTION ===\n{question.strip()}\n\n"
        f"=== MODEL RESPONSE ===\n{response.strip()}\n"
    )

def parse_score_block(text):
    keys = [
        "Consistency", "Factual Accuracy", "Completeness",
        "Citation Reliability", "Inference Justification"
    ]
    scores = []
    for key in keys:
        match = re.search(rf"{key}:\s*(0(?:\.\d+)?|1(?:\.0+)?)", text, re.IGNORECASE)
        scores.append(float(match.group(1)) if match else None)
    final = re.search(r"Final Score.*?:\s*(0(?:\.\d+)?|1(?:\.0+)?)", text, re.IGNORECASE)
    scores.append(float(final.group(1)) if final else None)
    return scores

def run_evaluation():
    try:
        with open(json_input_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook[sheet_name]
    except Exception as e:
        print(f" File load error: {e}")
        return

    driver = Driver(uc=True, headless=False)
    prompt_counter = 0
    start_time = time.time()

    try:
        driver.uc_open_with_reconnect(chatgpt_url, reconnect_time=30)
        print(" Log in manually to ChatGPT. Press ENTER when the chat interface is ready.")
        input()
#----------------------------------------------------------------------------------------------------------------------------------------#
        for i in range(2, 12):
            if i - 2 >= len(data):
                break

            entry = data[i - 2]
            question = entry.get("question")
            if not question:
                continue

            print(f"\n Row {i} → Processing 5 responses...")
            final_scores, bart_scores, entropy_scores, cosine_scores = [], [], [], []

            try:
                for j, key in enumerate(response_keys):
                    response = entry.get(key)
                    if not response:
                        continue

                    prompt_counter += 1
                    if prompt_counter >= PROMPT_LIMIT:
                        elapsed = time.time() - start_time
                        if elapsed < TIME_WINDOW:
                            wait_time = TIME_WINDOW - elapsed
                            print(f"Rate limit hit. Waiting {wait_time:.1f} seconds...")
                            time.sleep(wait_time)
                        start_time = time.time()
                        prompt_counter = 0

                    combined_prompt = format_prompt(question, response)

                    input_box = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div[contenteditable='true']"))
                    )
                    input_box.click()
                    time.sleep(3)
                    input_box.send_keys(Keys.CONTROL + "a")
                    input_box.send_keys(Keys.BACKSPACE)
                    pyperclip.copy(combined_prompt)
                    input_box.send_keys(Keys.CONTROL, 'v')
                    input_box.send_keys(Keys.ENTER)

                    WebDriverWait(driver, 60).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "div.markdown.prose.w-full.break-words"))
                    )
                    time.sleep(5)

                    response_divs = driver.find_elements(By.CSS_SELECTOR, "div.markdown.prose.w-full.break-words")
                    model_score_text = response_divs[-1].text.strip()
                    scores = parse_score_block(model_score_text)

                    if all(s is not None for s in scores):
                        final_scores.append(scores[5])
                        reference = f"{context_text.strip()} {question.strip()}"
                        bart = compute_bartscore(reference, response)
                        entropy = compute_semantic_entropy(response)
                        cosine = compute_cosine_similarity(reference, response)
                        bart_scores.append(bart)
                        entropy_scores.append(entropy)
                        cosine_scores.append(cosine)
                        print(f" R{i} | response_{j+1} → G-Eval: {scores[5]:.4f}, BART: {bart:.4f}, Entropy: {entropy:.4f}, Cosine: {cosine:.4f}")
                    else:
                        print(f" R{i} | response_{j+1} → Score parsing failed.")

                    time.sleep(WAIT_BETWEEN_RESPONSES)

                if final_scores:
                    g_eval = np.nanmean(final_scores).round(4)
                    bart_aggregate = np.nanmean(bart_scores).round(4) if bart_scores else None
                    entropy_aggregate = np.nanmean(entropy_scores).round(4) if entropy_scores else None
                    cosine_aggregate = np.nanmean(cosine_scores).round(4) if cosine_scores else None
                    all_scores = [g_eval, bart_aggregate, entropy_aggregate, cosine_aggregate]
                    for k, score in enumerate(all_scores):
                        sheet.cell(row=i, column=output_start_col + k, value=score)
                    print(f" R{i} → G-Eval and metrics written to Excel.")
                else:
                    sheet.cell(row=i, column=output_start_col, value="NO_VALID_SCORES")
                    print(f" R{i} → No valid scores collected.")

            except Exception as err:
                print(f" Error during row {i} processing: {err}")
                sheet.cell(row=i, column=output_start_col, value="ERROR")

    finally:
        workbook.save(excel_path)
        workbook.close()
        driver.quit()
        print(" DONE — All scores saved. ChatGPT session closed.")

if __name__ == "__main__":
    run_evaluation()