##########################################################################################################################

#---------------------------------- QA Generator ------------------------------------------------------------------------#

##############################################################################################################################


import json
import time
import io
import openpyxl
from contextlib import redirect_stdout
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from seleniumbase import Driver

qa_jsonl_path = r"C:\Users\XXXXXXXXX\Healthcare GPT Assessment\test.jsonl"
excel_path = r"C:\Users\XXXXXXXXX\Healthcare GPT Assessment\gpt_metadata_output_MedLLM_processed_0.xlsx"
json_output_path = r"C:\Users\XXXXXXXXX\Healthcare GPT Assessment\gpt_responses_bottom_250_1.json"
sheet_name = "Bottom 250"

login_url = "https://auth0.openai.com/u/login/password?state=hKFo2SBtbFFXQmhscnZVcWxtcnV6VFVWazhGanhUTHJFWi1NOaFur3VuaXZlcnNhbC1sb2dpbqN0aWTZIFZ5XzZROEVMYlZKbWdmbE1POU5PZzYyakVMaGZZenV0o2NpZNkgVGRKSWNiZTE2V29USHROOTVueXl3aDVFNHlPbzZJdEc"

model_column = 5
url_column = 6
start_row = 2
#end_row = 3  
repeat_count = 5
delay_between_turns = 20
PROMPT_LIMIT = 160
TIME_WINDOW = 3600

def load_first_question(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.loads(f.readline().strip())["question"]

def clean_response(text):
    if not text:
        return "NO_RESPONSE"
    if text.strip().lower().startswith("question:"):
        parts = text.split("Question:", 1)
        text = parts[-1].strip()

    closers = [
        "Let me know if you have any questions.",
        "Is there anything else I can help you with?",
        "I hope this helps.",
        "Please let me know if you'd like more detail.",
        "If you need more assistance, feel free to ask."
    ]
    for phrase in closers:
        text = text.replace(phrase, "").strip()

    lines = text.splitlines()
    seen = set()
    cleaned = []
    for line in lines:
        l = line.strip()
        if l and l not in seen:
            seen.add(l)
            cleaned.append(l)

    return "\n".join(cleaned).strip()

def run_evaluation():
    try:
        question_text = load_first_question(qa_jsonl_path)
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook[sheet_name]
        results = []
    except Exception as file_error:
        print(f" File load error: {file_error}")
        return

    driver = Driver(uc=True, headless=False)
    prompt_counter = 0
    start_time = time.time()

    try:
        driver.uc_open_with_reconnect(login_url, reconnect_time=30)
        print(" Please log in to ChatGPT manually. Press ENTER when ready.")
        driver.uc_gui_click_captcha()
        input()
#-----------------------------------------------------------------------------------------------------------------#
        for row in range(124, 125):
            model_name = sheet.cell(row=row, column=model_column).value
            gpt_url = sheet.cell(row=row, column=url_column).value
            if not gpt_url:
                continue

            print(f"\n Row {row} → {model_name} @ {gpt_url}")
            try:
                with io.StringIO() as buf, redirect_stdout(buf):
                    driver.get(gpt_url)

                WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "div#prompt-textarea"))
                )
                time.sleep(5)

                responses = []

                for i in range(repeat_count):
                    print(f" Sending question #{i+1}/5...")

                    prompt_counter += 1
                    if prompt_counter >= PROMPT_LIMIT:
                        elapsed = time.time() - start_time
                        if elapsed < TIME_WINDOW:
                            wait_time = TIME_WINDOW - elapsed
                            print(f"Rate limit hit. Waiting {wait_time:.1f} seconds...")
                            time.sleep(wait_time)
                        start_time = time.time()
                        prompt_counter = 0

                    try:
                        input_box = WebDriverWait(driver, 45).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "div[contenteditable='true']#prompt-textarea"))
                        )
                        input_box.click()
                        time.sleep(1.5)

                        driver.execute_script("""
                            const textarea = document.querySelector("div[contenteditable='true']#prompt-textarea");
                            const dataTransfer = new DataTransfer();
                            dataTransfer.setData('text', arguments[0]);
                            const event = new ClipboardEvent('paste', { clipboardData: dataTransfer, bubbles: true });
                            textarea.dispatchEvent(event);
                        """, question_text)

                        time.sleep(3.5)  

                        try:
                            WebDriverWait(driver, 20).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='send-button']"))
                            ).click()
                        except:
                            pass

                        WebDriverWait(driver, 60).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "div.markdown.prose.w-full.break-words"))
                        )
                        time.sleep(6)

                        response_divs = driver.find_elements(By.CSS_SELECTOR, "div.markdown.prose.w-full.break-words")
                        raw_text = response_divs[-1].text.strip() if response_divs else "NO_RESPONSE"
                        response_text = clean_response(raw_text)
                        responses.append(response_text)
                        print(f" Response {i+1} captured.")
                        time.sleep(delay_between_turns)

                    except Exception as e:
                        print(f" Failed on repetition {i+1}: {e}")
                        responses.append("ERROR")

                result_entry = {
                    "model": model_name,
                    "url": gpt_url,
                    "question": question_text
                }
                for i, r in enumerate(responses, 1):
                    result_entry[f"response_{i}"] = r

                results.append(result_entry)

            except Exception as err:
                print(f" GPT page error at row {row}: {err}")
                results.append({
                    "model": model_name or "UNKNOWN",
                    "url": gpt_url or "UNKNOWN",
                    "question": question_text,
                    **{f"response_{i}": "ERROR" for i in range(1, repeat_count + 1)}
                })

    finally:
        try:
            with open(json_output_path, "w", encoding="utf-8") as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
            print(f"\n All responses saved to:\n{json_output_path}")
        except Exception as save_error:
            print(f" JSON save failed: {save_error}")

        workbook.close()
        driver.quit()
        print(" DONE — Browser closed and workbook saved.")

if __name__ == "__main__":
    run_evaluation()